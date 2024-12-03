"""
VENTANA PARA LA FUNCIÓN TER
"""
from customtkinter import * 
import tkinter.messagebox
import openpyxl
import shutil
from _lib.get_files_path import get_file_paths
import xlwings as xw
from _lib.kill_excel_process import kill_excel_processes
from _lib.parse_coordinates import parse_coordinates
import time
import math
import PyPDF2
from _lib.progress_window import create_progress_window

def open_ter_window(menu_window,images_path):

   # Crear una nueva ventana
    window = CTkToplevel(menu_window)
    # Geometría
    width =410
    height = 200
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f'{width}x{height}+{x}+{y}')
    # Nombre de la ventana
    window.title("TER")
    # Ícono ventana
    window.after(201, lambda: window.iconbitmap(os.path.join(images_path, "ter.ico")))
    # Resizable
    window.resizable(False, False)
    window.grab_set()  # Esto hace que la ventana sea modal
    window.transient(menu_window)  # Para que esté asociada a la ventana principal
    window.focus_force()  # Enfoca la ventana modal

    #Variables para almacenar archivos
    file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
    input_path = ""

    #Función para adjuntar archivo del input
    def open_file_dialog(browse_entry):

        nonlocal file_content

        #ruta del archivo
        file_path = filedialog.askopenfilename()
        #extensión del archivo
        file_extension = os.path.splitext(file_path)[1]
        #nombre del archivo
        file_name = os.path.basename(file_path)  # Obtiene solo el nombre del archivo

        if file_extension != ".xlsx":
            tkinter.messagebox.showerror("Error","Formato de archivo no válido.")
            return
        else:
            browse_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            browse_entry.delete(0, "end")  # Limpia el contenido actual del campo
            browse_entry.insert(0, file_name)  # Inserta la ruta del archivo en el campo
            browse_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

            #Abre el archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            #Verifica que exista la información del proyecto, cliente y OS
            if sheet['B1'].value != None and sheet['B2'].value != None and sheet['B3'].value != None:
                file_content["OS"] = sheet['B1'].value
                file_content["cliente"] = sheet['B3'].value
                file_content["proyecto"] = sheet['B2'].value
            else:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
        
            #Última fila con información
            ultima_fila = 12
            for fila in range(12, sheet.max_row + 1):
                celda = sheet[f'A{fila}']
                if celda.value:  # Verifica si la celda tiene contenido
                    ultima_fila = fila

            if ultima_fila == 12:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
            
            #Obtiene los documentos a generar con sus respectivas coordenadas
            for row in sheet.iter_rows(min_row=12, max_row=ultima_fila, values_only=True):  # Ajusta según la tabla
                name = row[1]  
                position = row[4]
                geofonos =  row[19]  
                separacion =  row[20]  

                if name:
                    name_parts = name.split()
                    level = name_parts[0]  
                    group = name_parts[1]  
                    # Asegurarse de que la clave de nivel exista
                    if level not in file_content:
                        file_content[level] = {}

                    # Agregar el grupo y la posición
                    file_content[level][group] = position

                    if geofonos != None:
                        file_content[level]["geofonos"] = geofonos
                        file_content[level]["separacion"] = separacion


    #Función para adjuntar archivo del input
    def open_file_path(directory_entry, file_content):

        nonlocal input_path

        #ruta del archivo
        file_path = filedialog.askdirectory(title="Seleccionar un directorio")
        try:
            #contenido del directorio
            directory_content = os.listdir(file_path)
        except:
            tkinter.messagebox.showerror("Error", "La ruta no puede estar vacía.")
            return

        #Verifica que se haya insertado el input
        if file_content["cliente"] != "" and file_content["proyecto"] != "" and file_content["OS"] != "":
            
            files = list(file_content.keys())
            files.remove("cliente")
            files.remove("proyecto")
            files.remove("OS")

            #Verifica que en el directorio existan las carpetas correspondientes a los archivos
            for f in files:
                if f in directory_content:
                    
                    imagenes = [archivo for archivo in os.listdir(f'{file_path}/{f}') if archivo.lower().endswith(('.png', '.jpg'))]

                    if len(imagenes) == 1:

                        directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                        directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                        directory_entry.insert(0, file_path)  # Inserta la ruta del archivo en el campo
                        directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

                    else:

                        tkinter.messagebox.showerror("Error", "Sólo debe haber un archivo .png o .jpg en el directorio.")
                        return
                
                else:
                    tkinter.messagebox.showerror("Error", "La carpeta con los archivos de entrada no coincide con el input suministrado.")
                    directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                    directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                    directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite
                    return
        else:
            tkinter.messagebox.showerror("Error", "Debe seleccionar primero el archivo de entrada.")
            directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
            directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite
            return

    #Label seleccionar archivo de entrada
    label_archivo_entrada = CTkLabel(master=window,text="Archivo de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_archivo_entrada.place(x= 20,y=20)

    #Cajita para almacenar el nombre del archivo
    browse_entry = CTkEntry(master=window, width=140, font=('Gothic A1', 12), placeholder_text="Archivo",state=DISABLED)
    browse_entry.place(x=160, y=20)

    # Botón para buscar archivo
    browse_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_dialog(browse_entry)  
    )
    browse_button.place(x=310, y=23)

    #Label seleccionar archivos complementarios del input
    label_directory_input = CTkLabel(master=window,text="Directorio de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_directory_input.place(x= 20,y=80)

    #Cajita para almacenar la ruta
    directory_entry = CTkEntry(master=window, width=125, font=('Gothic A1', 12), placeholder_text="Directorio",state=DISABLED)
    directory_entry.place(x=175, y=80)

    # Botón para buscar ruta
    browse_path_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_path(directory_entry,file_content)  
    )
    browse_path_button.place(x=310, y=83)

    # Ejecutar 
    run_button = CTkButton(
        master=window,
        text="Generar reportes",
        width=100,
        height=30,
        command= lambda: ter_module(file_content,directory_entry.get()) 
    )
    run_button.place(x=(width - 100) // 2, y=140)
    


"""
EJECUTA EL MÓDULO DE TER
"""
def ter_module(file_content,inputs_path):

    #Verifica que se haya seleccionado la ruta de los archivos de entrada
    if inputs_path == "":
        tkinter.messagebox.showerror("Error", "Favor seleccionar la ruta de los archivos de entrada.")
        return
    
    #Verifica que se haya seleccionado el archivo de entrada
    if file_content["cliente"] == "" or file_content["proyecto"] == "" or file_content["OS"] == "":
        tkinter.messagebox.showerror("Error", "Favor cargar archivo de entrada.")
        return

    #Cierra todos los procesos de excel abiertos
    kill_excel_processes()
    time.sleep(2)

    #Templates path
    templates_path = get_file_paths("_templates")

    #Ruta de documentos
    documents_path = os.path.expanduser("~\\Documents")

    # Crea la carpeta para almacenar
    results_path = f'{documents_path}//GEOSTREAM//TER'
    if not os.path.exists(results_path):
        os.makedirs(results_path)
    else:
        # Elimina los archivos adentro si la carpeta ya existe
        for elemento in os.listdir(results_path):
            ruta_elemento = os.path.join(results_path, elemento)
            if os.path.isdir(ruta_elemento):
                shutil.rmtree(ruta_elemento)  
            else:
                os.remove(ruta_elemento)

    #Crea la ventana de progreso
    ventana_progreso, barra_progreso, texto_progreso = create_progress_window("TER - Generando memorias","logo.ico",f"0/{len(list(file_content.keys()))-3}")

    #Itera por los archivos
    cont = 1
    for key in list(file_content.keys()):
        if key != "cliente" and key != "proyecto" and key != "OS":

            #Copia el template
            os.mkdir(f'{documents_path}//GEOSTREAM//TER//{key}')
            shutil.copy(f'{templates_path}//template ter.xlsm', f'{documents_path}///GEOSTREAM//TER//{key}//template ter.xlsm')

            #Abre el template
            app = xw.App(visible=False)  
            app.display_alerts = False 
            app.screen_updating = False
            libro = xw.Book(f'{documents_path}//GEOSTREAM//TER//{key}//template ter.xlsm', update_links=True)

            #Modifica hoja A1
            modificar_hoja_A1(libro,file_content,key)

            #Modificar hoja fotos: Espectro G01, Espectro G12, Espectro G24, Inversión Linea G01, Inversión Linea G12, Inversión Linea G24, Fotos Linea
            modificar_hoja_fotos(libro,key,inputs_path)

            #Imprimir en PDF la memoria
            save_to_pdf(libro, key, documents_path )

            # Guardar los cambios
            libro.save()

            #Cerrar
            libro.close()
            app.quit()

            #Actualizar barra de progreso
            barra_progreso.set(cont / (len(list(file_content.keys()))-3))  # Actualizar progreso
            texto_progreso.set(f"{cont}/{(len(list(file_content.keys()))-3)}")
            ventana_progreso.update_idletasks()  # Forzar actualización de la ventana
            cont+=1

    #Compilar versión final
    merger = PyPDF2.PdfMerger()
    for key in list(file_content.keys()):
        if key != "cliente" and key != "proyecto" and key != "OS":
            merger.append(f'{documents_path}//GEOSTREAM//TER//{key}//{key}.pdf')

    # Guardar el PDF combinado en un archivo
    merger.write(f'{documents_path}//GEOSTREAM//TER//combinado.pdf')
    merger.close()

    ventana_progreso.destroy()
    tkinter.messagebox.showinfo("Info","Memorias generadas. Revisar en Documentos.")

def modificar_hoja_A1(libro,file_content,key):

    hoja = libro.sheets[0]  
    hoja["B3"].value = file_content["proyecto"]
    hoja["E3"].value = file_content["OS"]
    hoja["E4"].value = file_content["cliente"]
    hoja["B6"].value = key
    hoja["B10"].value = file_content[key]["geofonos"]
    hoja["B11"].value = file_content[key]["separacion"]

    #Tratamiendo de coordenadas
    latitud_start, longitud_start = parse_coordinates(file_content[key]["G01"])
    latitud_end, longitud_end = parse_coordinates(file_content[key]["G24"])

    #Escribe las coordenadas
    hoja["E9"].value = longitud_start
    hoja["F9"].value = latitud_start
    hoja["E10"].value = longitud_end
    hoja["F10"].value = latitud_end

def modificar_hoja_fotos(libro,key,inputs_path):

    #Parsea el input_path
    inputs_path = inputs_path.replace("/", "\\")

    #Obtiene la hoja
    hoja = libro.sheets["TER01"]

    #Nombre de la figura
    img_name = os.listdir(f"{inputs_path}\{key}")[0]
        
    #Borra y pega la imagen
    hoja.pictures.add(f"{inputs_path}\{key}\{img_name}",
                    top = hoja.range("B9").top,
                    width = 720)
    
def save_to_pdf(libro, key, documents_path ):

    #Obtiene la hoja
    hoja = libro.sheets["TER01"]

    #Guarda el pdf
    output_pdf_path = f'{documents_path}//GEOSTREAM//TER//{key}//{key}.pdf'

    hoja.api.ExportAsFixedFormat(
        Type=0,  # 0 es para PDF
        Filename=output_pdf_path,
        Quality=0,  # Calidad estándar
        IncludeDocProperties=True,
        IgnorePrintAreas=False,  # Respetar áreas de impresión
        OpenAfterPublish=False
    )  

