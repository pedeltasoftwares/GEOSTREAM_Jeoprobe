"""
VENTANA PARA LA FUNCIÓN SPAC
"""
from customtkinter import * 
import tkinter.messagebox
import openpyxl
import shutil
from _lib.get_files_path import get_file_paths
import xlwings as xw
from _lib.kill_excel_process import kill_excel_processes
from _lib.parse_coordinates import parse_coordinates
import time
import math
import PyPDF2
from _lib.progress_window import create_progress_window
from _lib.set_cell_border import set_border

def open_spac_window(menu_window,images_path):

    # Crear una nueva ventana
    window = CTkToplevel(menu_window)
    # Geometría
    width =410
    height = 200
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f'{width}x{height}+{x}+{y}')
    # Nombre de la ventana
    window.title("SPAC")
    # Ícono ventana
    window.after(201, lambda: window.iconbitmap(os.path.join(images_path, "spac.ico")))
    # Resizable
    window.resizable(False, False)
    window.grab_set()  # Esto hace que la ventana sea modal
    window.transient(menu_window)  # Para que esté asociada a la ventana principal
    window.focus_force()  # Enfoca la ventana modal

    #Variables para almacenar archivos
    file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
    input_path = ""

    #Función para adjuntar archivo del input
    def open_file_dialog(browse_entry):

        nonlocal file_content

        #ruta del archivo
        file_path = filedialog.askopenfilename()
        #extensión del archivo
        file_extension = os.path.splitext(file_path)[1]
        #nombre del archivo
        file_name = os.path.basename(file_path)  # Obtiene solo el nombre del archivo

        if file_extension != ".xlsx":
            tkinter.messagebox.showerror("Error","Formato de archivo no válido.")
            return
        else:
            browse_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            browse_entry.delete(0, "end")  # Limpia el contenido actual del campo
            browse_entry.insert(0, file_name)  # Inserta la ruta del archivo en el campo
            browse_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

            #Abre el archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            #Verifica que exista la información del proyecto, cliente y OS
            if sheet['B1'].value != None and sheet['B2'].value != None and sheet['B3'].value != None:
                file_content["OS"] = sheet['B1'].value
                file_content["cliente"] = sheet['B3'].value
                file_content["proyecto"] = sheet['B2'].value
            else:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
        
            #Última fila con información
            ultima_fila = 12
            for fila in range(12, sheet.max_row + 1):
                celda = sheet[f'A{fila}']
                if celda.value:  # Verifica si la celda tiene contenido
                    ultima_fila = fila

            if ultima_fila == 12:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "OS": ""
                    }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
            
            #Obtiene los documentos a generar con sus respectivas coordenadas
            for row in sheet.iter_rows(min_row=12, max_row=ultima_fila, values_only=True):  # Ajusta según la tabla
                name = row[1]  
                position = row[4]
                geofonos =  row[19]  
                separacion =  row[20]  

                if name:
                    name_parts = name.split()
                    level = name_parts[0]  
                    group = name_parts[1]  
                    # Asegurarse de que la clave de nivel exista
                    if level not in file_content:
                        file_content[level] = {}

                    # Agregar el grupo y la posición
                    file_content[level][group] = position

                    if geofonos != None:
                        file_content[level]["geofonos"] = geofonos
                        file_content[level]["separacion"] = separacion


    #Función para adjuntar archivo del input
    def open_file_path(directory_entry, file_content):

        nonlocal input_path

        #ruta del archivo
        file_path = filedialog.askdirectory(title="Seleccionar un directorio")
        try:
            #contenido del directorio
            directory_content = os.listdir(file_path)
        except:
            tkinter.messagebox.showerror("Error", "La ruta no puede estar vacía.")
            return

        #Archivos necesarios para completar la verificación
        verify_files = ['espectro.png', 'SPAC.fv', 'inversión.png', 'FIN.jpg', 'INI.jpg']

        #Verifica que se haya insertado el input
        if file_content["cliente"] != "" and file_content["proyecto"] != "" and file_content["OS"] != "":
            
            files = list(file_content.keys())
            files.remove("cliente")
            files.remove("proyecto")
            files.remove("OS")

            #Verifica que en el directorio existan las carpetas correspondientes a los archivos
            for f in files:
                if f in directory_content:
                    try:
                        assert set(os.listdir(f'{file_path}//{f}')) == set(verify_files), "Los elementos no coinciden"

                        directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                        directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                        directory_entry.insert(0, file_path)  # Inserta la ruta del archivo en el campo
                        directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

                    except AssertionError as e:

                        directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                        directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                        directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

                        tkinter.messagebox.showerror("Error", "No están todos los archivos necesarios para ejecutar el aplicativo.")
                        return
                
                else:
                    tkinter.messagebox.showerror("Error", "La carpeta con los archivos de entrada no coincide con el input suministrado.")
                    directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                    directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                    directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite
                    return
        else:
            tkinter.messagebox.showerror("Error", "Debe seleccionar primero el archivo de entrada.")
            directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
            directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite
            return

    #Label seleccionar archivo de entrada
    label_archivo_entrada = CTkLabel(master=window,text="Archivo de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_archivo_entrada.place(x= 20,y=20)

    #Cajita para almacenar el nombre del archivo
    browse_entry = CTkEntry(master=window, width=140, font=('Gothic A1', 12), placeholder_text="Archivo",state=DISABLED)
    browse_entry.place(x=160, y=20)

    # Botón para buscar archivo
    browse_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_dialog(browse_entry)  
    )
    browse_button.place(x=310, y=23)

    #Label seleccionar archivos complementarios del input
    label_directory_input = CTkLabel(master=window,text="Directorio de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_directory_input.place(x= 20,y=80)

    #Cajita para almacenar la ruta
    directory_entry = CTkEntry(master=window, width=125, font=('Gothic A1', 12), placeholder_text="Directorio",state=DISABLED)
    directory_entry.place(x=175, y=80)

    # Botón para buscar ruta
    browse_path_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_path(directory_entry,file_content)  
    )
    browse_path_button.place(x=310, y=83)

    # Ejecutar 
    run_button = CTkButton(
        master=window,
        text="Generar reportes",
        width=100,
        height=30,
        command= lambda: spac_module(file_content,directory_entry.get(),window) 
    )
    run_button.place(x=(width - 100) // 2, y=140)
    


"""
EJECUTA EL MÓDULO DE SPAC
"""
def spac_module(file_content,inputs_path,window):

    #Verifica que se haya seleccionado la ruta de lso archivos de entrada
    if inputs_path == "":
        tkinter.messagebox.showerror("Error", "Favor seleccionar la ruta de los archivos de entrada.")
        return
    
    #Verifica que se haya seleccionado el archivo de entrada
    if file_content["cliente"] == "" or file_content["proyecto"] == "" or file_content["OS"] == "":
        tkinter.messagebox.showerror("Error", "Favor cargar archivo de entrada.")
        return

    #Cierra todos los procesos de excel abiertos
    kill_excel_processes()
    time.sleep(2)

    #Templates path
    templates_path = get_file_paths("_templates")

    #Ruta de documentos
    documents_path = os.path.expanduser("~\\Documents")

    # Crea la carpeta para almacenar
    results_path = f'{documents_path}//GEOSTREAM//SPAC'
    if not os.path.exists(results_path):
        os.makedirs(results_path)
    else:
        # Elimina los archivos adentro si la carpeta ya existe
        for elemento in os.listdir(results_path):
            ruta_elemento = os.path.join(results_path, elemento)
            if os.path.isdir(ruta_elemento):
                shutil.rmtree(ruta_elemento)  
            else:
                os.remove(ruta_elemento)

    #Crea la ventana de progreso
    ventana_progreso, barra_progreso, texto_progreso = create_progress_window("SPAC - Generando memorias","logo.ico",f"0/{len(list(file_content.keys()))-3}")

    #Itera por los archivos
    cont = 1
    for key in list(file_content.keys()):
        if key != "cliente" and key != "proyecto" and key != "OS":

            #Copia el template
            os.mkdir(f'{documents_path}//GEOSTREAM//SPAC//{key}')
            shutil.copy(f'{templates_path}//template spac.xlsm', f'{documents_path}///GEOSTREAM//SPAC//{key}//template spac.xlsm')

            #Abre el template
            app = xw.App(visible=False)  
            app.display_alerts = False 
            app.screen_updating = False
            libro = xw.Book(f'{documents_path}//GEOSTREAM//SPAC//{key}//template spac.xlsm', update_links=True)

            #Modifica hoja A1
            modificar_hoja_A1(libro,file_content,key)

            #Modificar hoja A2
            modificar_hoja_A2(libro,key,inputs_path)

            #Modifica la hoja de modulos elastivos
            modificar_modulos_elasticos(libro,key)

            #Modificar hoja fotos: Espectro G01, Inversión Linea G01, Fotos Linea
            modificar_hoja_fotos(libro,key,inputs_path)

            #Imprimir en PDF la memoria
            save_to_pdf(libro, key, documents_path )

            # Guardar los cambios
            libro.save()

            #Cerrar
            libro.close()
            app.quit()

            #Actualizar barra de progreso
            barra_progreso.set(cont / (len(list(file_content.keys()))-3))  # Actualizar progreso
            texto_progreso.set(f"{cont}/{(len(list(file_content.keys()))-3)}")
            ventana_progreso.update_idletasks()  # Forzar actualización de la ventana
            cont+=1

    #Compilar versión final
    merger = PyPDF2.PdfMerger()
    for key in list(file_content.keys()):
        if key != "cliente" and key != "proyecto" and key != "OS":
            merger.append(f'{documents_path}//GEOSTREAM//SPAC//{key}//{key}_combinado.pdf')

    # Guardar el PDF combinado en un archivo
    merger.write(f'{documents_path}//GEOSTREAM//SPAC//combinado.pdf')
    merger.close()

    ventana_progreso.destroy()
    tkinter.messagebox.showinfo("Info","Memorias generadas. Revisar en Documentos.")
    window.destroy()

def modificar_hoja_A1(libro,file_content,key):

    hoja = libro.sheets[0]  
    hoja["B3"].value = file_content["proyecto"]
    hoja["E3"].value = file_content["OS"]
    hoja["E4"].value = file_content["cliente"]
    hoja["B6"].value = key
    hoja["B10"].value = file_content[key]["geofonos"]
    hoja["B11"].value = file_content[key]["separacion"]

    #Tratamiendo de coordenadas
    latitud_start, longitud_start = parse_coordinates(file_content[key]["G01"])
    latitud_end, longitud_end = parse_coordinates(file_content[key]["G24"])

    #Escribe las coordenadas
    hoja["E9"].value = longitud_start
    hoja["F9"].value = latitud_start
    hoja["E10"].value = longitud_end
    hoja["F10"].value = latitud_end

def modificar_hoja_A2(libro,key,inputs_path):

    #selecciona la hoja
    hoja = libro.sheets["A2"]

    #Elimina las filas existentes
    rango = hoja.range('A6:G50')
    rango.value = None
        
    #abre los el archivo .fv y los lee
    name_file = 'SPAC.fv'
    with open(f'{inputs_path}//{key}//{name_file}','r') as f:
        lines = f.readlines()
    index = []
    for line in lines:
        if "MODEL_INVERTED: " in line:
            index.append(lines.index(line))
        
        if "MISFIT_RMS_NRMS: " in line:
            index.append(lines.index(line))

    #Toma los valores
    col_1, col_2 = [],[]
    for i in range(index[0]+1, index[1]):
        col_1.append(float(lines[i].split("\n")[0].split(" ")[0]))
        col_2.append(float(lines[i].split("\n")[0].split(" ")[2]))
    
    #copia los nuevos valores
    fila = 6
    for i in range(len(col_1)):
        #Espesor
        hoja[f"A{fila}"].value = col_1[i]
        #Vs
        hoja[f"B{fila}"].value = col_2[i]
        #profundidad
        if i != 0:
            hoja[f"C{fila}"].value = hoja[f"C{fila-1}"].value + col_1[i]
        else:
            hoja[f"C{fila}"].value = col_1[i]
        
        fila+=1

    #Modelos VS
    hoja["D5"].value = 0
    hoja["E5"].value = hoja["B5"].value
    fila = 6
    for i in range(len(col_1)):
       
        hoja[f"D{fila}"].value = hoja[f"C{i+6}"].value
        hoja[f"E{fila}"].value = hoja[f"B{i+6}"].value

        hoja[f"D{fila+1}"].value = hoja[f"C{i+6}"].value
        hoja[f"E{fila+1}"].value = hoja[f"B{i+7}"].value

        #Poisson
        nu_1 = 1.8945 * hoja[f"E{fila}"].value ** (-0.325)
        if nu_1 > 0.495:
            hoja[f"F{fila}"].value =  0.495
        else:
            hoja[f"F{fila}"].value = nu_1

        try:
            nu_2 = 1.8945 * hoja[f"E{fila + 1}"].value ** (-0.325)
            if nu_2 > 0.495:
                hoja[f"F{fila + 1}"].value =  0.495
            else:
                hoja[f"F{fila + 1}"].value = nu_2
        except:
            pass
        
        hoja[f"G{fila}"].value = ( ( ( 2 - 2 * hoja[f"F{fila}"].value ) / ( 1 - 2 * hoja[f"F{fila}"].value ) ) ** (0.5) ) * hoja[f"E{fila}"].value

        try:
            hoja[f"G{fila + 1}"].value = ( ( ( 2 - 2 * hoja[f"F{fila + 1}"].value ) / ( 1 - 2 * hoja[f"F{fila + 1}"].value ) ) ** (0.5) ) * hoja[f"E{fila + 1}"].value
        except:
            pass
        
        if i == len(col_1) - 1:
            hoja[f"D{fila+1}"].value = None

        fila+=2

    hoja["F5"].value = hoja["F6"].value
    hoja["G5"].value = hoja["G6"].value

def modificar_modulos_elasticos(libro,key):

    #Obtiene la hoja de módulos elasticops
    hoja_modulos_elasticos = libro.sheets["Módulos elásticos"]

    rango = hoja_modulos_elasticos.range('A12:N28')
    rango.value = None

    #Parámetros
    a = 19430
    b = 1.960784314
    efic = 1

    #Obtiene la hoja de A2
    hoja=libro.sheets["A2"]

    #Última fila con información
    ultima_fila = hoja.api.UsedRange.Rows.Count 
    for fila in range(5, ultima_fila + 1):
        celda = hoja[f'D{fila}']
        if celda.value:  # Verifica si la celda tiene contenido
            ultima_fila = fila

    #Columna para copiar el Vs en la hoja de modulos elasticos
    col_prof = "A" 
    col_Vs = "B" 
    #Copia la información
    for fila in range(5,ultima_fila+1):
        #profundidad
        hoja_modulos_elasticos[f'{col_prof}{fila+7}'].value = hoja[f"D{fila}"].value
        #Vs
        hoja_modulos_elasticos[f'{col_Vs}{fila+7}'].value = hoja[f"E{fila}"].value

    #Modifica la tabla del perfil promedio
    ultima_fila = hoja_modulos_elasticos.api.UsedRange.Rows.Count 
    for fila in range(12, ultima_fila + 1):
        celda = hoja_modulos_elasticos[f'A{fila}']
        if celda.value:  # Verifica si la celda tiene contenido
            ultima_fila = fila

    for fila in range(12,ultima_fila+1):

        #Promedio de profundidad
        hoja_modulos_elasticos[f"I{fila}"].value = round(  hoja_modulos_elasticos[f"A{fila}"].value, 1)
        hoja_modulos_elasticos.range(f"I{fila}").api.HorizontalAlignment = -4108
        set_border(hoja_modulos_elasticos,"I",fila)

        #Promedio de vs
        hoja_modulos_elasticos[f"J{fila}"].value = round( hoja_modulos_elasticos[f"B{fila}"].value, 1) 
        hoja_modulos_elasticos.range(f"J{fila}").api.HorizontalAlignment = -4108
        set_border(hoja_modulos_elasticos,"J",fila)
        
        if fila != 12:
            #Gamma saturado
            hoja_modulos_elasticos[f"L{fila}"].value = round( 8.32 * math.log10( hoja_modulos_elasticos[f"J{fila}"].value )  - 
                                                                1.61 * math.log10(hoja_modulos_elasticos[f"I{fila}"].value ), 1)
            hoja_modulos_elasticos.range(f"L{fila}").api.HorizontalAlignment = -4108
            set_border(hoja_modulos_elasticos,"L",fila)
            
            #Go
            hoja_modulos_elasticos[f"M{fila}"].value = round(  (hoja_modulos_elasticos[f"L{fila}"].value / 10) * hoja_modulos_elasticos[f"J{fila}"].value  ** 2, 0)
            hoja_modulos_elasticos.range(f"M{fila}").api.HorizontalAlignment = -4108
            set_border(hoja_modulos_elasticos,"M",fila)

            #Eo
            hoja_modulos_elasticos[f"N{fila}"].value = round( 2 * hoja_modulos_elasticos[f"M{fila}"].value * ( 1 + 0.33), 0)
            hoja_modulos_elasticos.range(f"N{fila}").api.HorizontalAlignment = -4108
            set_border(hoja_modulos_elasticos,"N",fila)

            #Nequiv
            if round(( (hoja_modulos_elasticos[f"M{fila}"].value / a) ** b ) * efic ,0) < 80:

                hoja_modulos_elasticos[f"K{fila}"].value = round(( (hoja_modulos_elasticos[f"M{fila}"].value / a) ** b ) * efic ,0)
                hoja_modulos_elasticos.range(f"K{fila}").api.HorizontalAlignment = -4108
                set_border(hoja_modulos_elasticos,"K",fila)

            else:
                hoja_modulos_elasticos[f"K{fila}"].value = "RECHAZO"
                hoja_modulos_elasticos.range(f"K{fila}").api.HorizontalAlignment = -4108
                set_border(hoja_modulos_elasticos,"K",fila)

        
    #Completa la primera linea
    hoja_modulos_elasticos["L12"].value = hoja_modulos_elasticos["L13"].value 
    hoja_modulos_elasticos["M12"].value = round(  (hoja_modulos_elasticos["L12"].value / 10) * hoja_modulos_elasticos["J12"].value  ** 2, 0)
    set_border(hoja_modulos_elasticos,"M",12)  
    hoja_modulos_elasticos["N12"].value = round( 2 * hoja_modulos_elasticos["M12"].value * ( 1 + 0.33), 0)
    set_border(hoja_modulos_elasticos,"N",12)  
    if round(( (hoja_modulos_elasticos["M12"].value / a) ** b ) * efic ,0) < 80:
        hoja_modulos_elasticos["K12"].value = round(( (hoja_modulos_elasticos["M12"].value / a) ** b ) * efic ,0)
    else:
        hoja_modulos_elasticos["K12"].value = "RECHAZO"
    set_border(hoja_modulos_elasticos,"K",12)   

    #Calcula el Vs
    results = []
    for fila in range(13, ultima_fila + 1,2):

        if hoja_modulos_elasticos[f"I{fila}"].value < 30:

            if fila == 12:
                results.append( round (hoja_modulos_elasticos[f"I{fila}"].value / hoja_modulos_elasticos[f"J{fila}"].value,4))
            else:
                results.append( round ( (hoja_modulos_elasticos[f"I{fila}"].value - hoja_modulos_elasticos[f"I{fila - 1}"].value)/ hoja_modulos_elasticos[f"J{fila}"].value,4))
        else:
            results.append( round ( (30 - hoja_modulos_elasticos[f"I{fila - 1}"].value)/ hoja_modulos_elasticos[f"J{fila}"].value,4))
            break

    #Tipo de suelo
    Vs_30 = round(30 / sum(results),0)
    if Vs_30 > 1500:
        tipo_suelo = "A"
    elif Vs_30 > 760 and Vs_30<= 1500:
        tipo_suelo = "B"
    elif Vs_30 > 360 and Vs_30<= 760:
        tipo_suelo = "C"
    elif Vs_30 > 180 and Vs_30<= 360:
        tipo_suelo = "D"
    else:
        "E/F"

    #Escribe el rotulo de Vs y el tipo de suelo
    hoja_modulos_elasticos[f"K{ultima_fila + 3}"].value = "Vs30 (m/s)"
    hoja_modulos_elasticos[f"K{ultima_fila + 3}"].font.bold = True
    hoja_modulos_elasticos.range(f"K{ultima_fila + 3}").characters[2:5].api.Font.Subscript = True
    hoja_modulos_elasticos[f"L{ultima_fila + 3}"].value = Vs_30
    set_border(hoja_modulos_elasticos,"K",ultima_fila + 3)
    set_border(hoja_modulos_elasticos,"L",ultima_fila + 3)   

    hoja_modulos_elasticos[f"K{ultima_fila + 4}"].value = "Tipo de suelo"
    hoja_modulos_elasticos[f"K{ultima_fila + 4}"].font.bold = True
    hoja_modulos_elasticos[f"L{ultima_fila + 4}"].value = tipo_suelo
    hoja_modulos_elasticos.range(f"L{ultima_fila + 4}").api.HorizontalAlignment = xw.constants.HAlign.xlHAlignLeft
    set_border(hoja_modulos_elasticos,"K",ultima_fila + 4)
    set_border(hoja_modulos_elasticos,"L",ultima_fila + 4)   

    #Modifica el eje Y de la figur y el eje X se las series
    chart = hoja_modulos_elasticos.charts[0] 

    # Actualizar la serie de datos de la gráfica
    num_series = chart.api[1].SeriesCollection().Count  
    for i in range(1, num_series + 1):  
        serie = chart.api[1].SeriesCollection(i)

        if "PERFIL PROMEDIO" not in serie.Name:
            columna_y = chr(65 + (i - 1) * 2) 
            columna_x = chr(66 + (i - 1) * 2)  
        else:
            columna_y = "I"
            columna_x = "J"

        serie.XValues = hoja_modulos_elasticos.range(f"{columna_x}12:{columna_x}{ultima_fila}").value
        serie.Values = hoja_modulos_elasticos.range(f"{columna_y}12:{columna_y}{ultima_fila}").value

    #Max del eje Y
    chart.api[1].Axes(2).MaximumScale = round(hoja_modulos_elasticos[f"A{ultima_fila}"].value,0)
    chart.api[1].Axes(2).MinimumScale = 0

def modificar_hoja_fotos(libro,key,inputs_path):

    #Parsea el input_path
    inputs_path = inputs_path.replace("/", "\\")

    for sheet_name in ["Espectro G01","Inversión Linea G01"]:

        #Obtiene la hoja
        hoja = libro.sheets[sheet_name]

        if "Espectro G" in sheet_name:
            img_name = "espectro.png"
            range_insert = "A10"
            width = "730"
        
        elif "Inversión Linea G" in sheet_name:
            img_name = "inversión.png"
            range_insert = "A10"
            width = "730"
    
        #Borra y pega la imagen
        hoja.pictures.add(f"{inputs_path}\{key}\{img_name}",
                        top = hoja.range(range_insert).top,
                        width = width)
    
    #Arregla la hoja de Fotos Linea
    hoja = libro.sheets["Fotos Linea"]

    #Obtiene las formas
    hoja.pictures.add(f"{inputs_path}\{key}\INI.jpg",
                    top = hoja.range("A9").top,
                    width = '290')
            
    hoja.pictures.add(f"{inputs_path}\{key}\FIN.jpg",
                    left = hoja.range("G9").left,
                    top = hoja.range("G9").top,
                    width = '290')

def save_to_pdf(libro, key, documents_path ):

    #hojas
    sheet_names = ["Módulos elásticos", "Espectro G01", "Inversión Linea G01","Fotos Linea"]
    for sheet_name in sheet_names:

        #Obtiene la hoja
        hoja = libro.sheets[sheet_name]

        #Guarda el pdf
        output_pdf_path = f'{documents_path}//GEOSTREAM//SPAC//{key}//{sheet_name}.pdf'

        hoja.api.ExportAsFixedFormat(
            Type=0,  # 0 es para PDF
            Filename=output_pdf_path,
            Quality=0,  # Calidad estándar
            IncludeDocProperties=True,
            IgnorePrintAreas=False,  # Respetar áreas de impresión
            OpenAfterPublish=False
        )   

    #Compila en uno solo
    merger = PyPDF2.PdfMerger()
    for pdf in sheet_names:
        merger.append(f'{documents_path}//GEOSTREAM//SPAC//{key}//{pdf}.pdf')

    # Guardar el PDF combinado en un archivo
    merger.write(f'{documents_path}//GEOSTREAM//SPAC//{key}//{key}_combinado.pdf')
    merger.close()

    for pdf in sheet_names:
        os.remove(f'{documents_path}//GEOSTREAM//SPAC//{key}//{pdf}.pdf')

